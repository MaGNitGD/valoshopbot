import asyncio
import logging
import datetime
import random
import aiogram
import openpyxl
import keyboards as kb
import database as db
import token

from aiogram import Bot, Dispatcher, types, F
from aiogram.utils.keyboard import ReplyKeyboardBuilder, InlineKeyboardBuilder
from aiogram.filters import Filter
from aiogram.types import Message
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup


f = open('skins.txt', 'r')
allSkins = f.read().split(', ')
#print(allSkins)
f.close()

allWeap = ["ke", "cc", "sy", "fy", "gt", "sf", "by", "ml", "sr", "as", "se", "je", "bg", "gn", "pm", "vl", "on", "or"]
userEquipCells = ['Bke', 'Ccc', 'Dsy', 'Efy', 'Fgt', 'Gsf', 'Hby', 'Iml', 'Jsr', 'Kas', 'Lse', 'Mje', 'Nbg', 'Ogn', 'Ppm', 'Qvl', 'Ron', 'Sor']

API_TOKEN = token.bot_token #токен бота
logging.basicConfig(level=logging.INFO) #сбор логов (куда?)

bot = Bot(token=API_TOKEN)
disp = Dispatcher(bot=bot)

columns = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

def findCell(value, column, raw, bdin, uinfin,):
    bd = openpyxl.load_workbook(bdin)  # открываю бд
    uinf = bd[uinfin]  # выбираю лист usersInfo
    i = raw  # образ raw в цикле
    currentCell = column + str(raw) # определяю проверяему ячейку
    currentCellValue = uinf[currentCell].value # определяю значение проверяемой ячейки
    none = False # пустая ли ячейка?

    x = 0  # отключение while
    while x != 1:
        raw = str(i)  # преобразую i в номер строки
        currentCell = (column+ str(raw)) # определяю проверяемую ячейку
        currentCellValue = uinf[currentCell].value  # определяю значение проверяемой ячейки
        if currentCellValue == value: # если ячейка найдена
            x = 1 # отключение while
            bd.save(bdin)
            return [currentCell, currentCellValue] # координаты ячейки и её значение
        elif uinf[currentCell].value is None: # если ячейки нет
            none = True
            bd.save(bdin)
            return [currentCell, none] # координаты ячеёки и ин-я, что она пуста
        else:
            i += 1

def setDefaultSkins(botId):
    botidCell = findCell(botId, 'A', 1, 'users.xlsx', 'usersInfo')
    bd = openpyxl.load_workbook('users.xlsx')
    ueqiup = bd['usersEquipped']
    ueqiup['B' + str(int(botidCell[0][1]) - 1)] = 'ke0'
    ueqiup['C' + str(int(botidCell[0][1]) - 1)] = 'cc0'
    ueqiup['D' + str(int(botidCell[0][1]) - 1)] = 'sy0'
    ueqiup['E' + str(int(botidCell[0][1]) - 1)] = 'fy0'
    ueqiup['F' + str(int(botidCell[0][1]) - 1)] = 'gt0'
    ueqiup['G' + str(int(botidCell[0][1]) - 1)] = 'sf0'
    ueqiup['H' + str(int(botidCell[0][1]) - 1)] = 'by0'
    ueqiup['I' + str(int(botidCell[0][1]) - 1)] = 'ml0'
    ueqiup['J' + str(int(botidCell[0][1]) - 1)] = 'sr0'
    ueqiup['K' + str(int(botidCell[0][1]) - 1)] = 'as0'
    ueqiup['L' + str(int(botidCell[0][1]) - 1)] = 'se0'
    ueqiup['M' + str(int(botidCell[0][1]) - 1)] = 'je0'
    ueqiup['N' + str(int(botidCell[0][1]) - 1)] = 'bg0'
    ueqiup['O' + str(int(botidCell[0][1]) - 1)] = 'gn0'
    ueqiup['P' + str(int(botidCell[0][1]) - 1)] = 'pm0'
    ueqiup['Q' + str(int(botidCell[0][1]) - 1)] = 'vl0'
    ueqiup['R' + str(int(botidCell[0][1]) - 1)] = 'on0'
    ueqiup['S' + str(int(botidCell[0][1]) - 1)] = 'or0'
    bd.save('users.xlsx')

def getUserStats(botId):
    botIdCell = findCell(botId, 'A', 2, 'users.xlsx', 'usersEquipped')
    ubd = openpyxl.load_workbook('users.xlsx')
    sbd = openpyxl.load_workbook('skins.xlsx')
    ueqip = ubd['usersEquipped']
    skins = sbd['skins']

    ke = skins['D' + findCell(ueqip['B' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    bg = skins['D' + findCell(ueqip['N' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    gn = skins['D' + findCell(ueqip['O' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    hp = ke + bg + (ke + bg) / 100 * gn

    fy = skins['D' + findCell(ueqip['E' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    ass = skins['D' + findCell(ueqip['K' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    on = skins['D' + findCell(ueqip['R' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    armor = fy + ass + (fy + ass) / 100 * on

    cc = skins['D' + findCell(ueqip['C' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    sr = skins['D' + findCell(ueqip['J' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    vl = skins['D' + findCell(ueqip['Q' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    damage = cc + sr + (cc + sr) / 100 * vl

    sf = skins['D' + findCell(ueqip['G' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    ml = skins['D' + findCell(ueqip['I' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    orr = skins['D' + findCell(ueqip['S' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    accuracy = sf + orr + (sf + orr) / 100 * ml

    gt = skins['D' + findCell(ueqip['F' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    se = skins['D' + findCell(ueqip['L' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    pm = skins['D' + findCell(ueqip['P' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    hs = gt + pm + (gt + pm) / 100 * se

    sy = skins['D' + findCell(ueqip['D' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    by = skins['D' + findCell(ueqip['H' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    je = skins['D' + findCell(ueqip['M' + botIdCell[0][1]].value, 'C', 1, 'skins.xlsx', 'skins')[0][1:]].value
    dodge = sy + je + (sy + je)/100*by

    return hp, armor, damage, accuracy, hs, dodge

def isRegistered(botId):
    bd = openpyxl.load_workbook('users.xlsx')  # открываю бд
    uinf = bd['usersInfo']  # выбираю лист usersInfo
    botid = findCell(botId, 'B', 2, 'users.xlsx', 'usersInfo')  # поиск юзера в базе
    if botid[1] == True:  # если юзер не зареган
        return False
    else:
        return True

def generateRandomStore():
    sbd = openpyxl.load_workbook('skins.xlsx')
    skins = sbd['skins']
    skinsCount = int(findCell(None, 'C', 1, 'skins.xlsx', 'skins')[0][1:]) - 2
    a = skins['C' + str(random.randint(2, skinsCount))].value
    b = skins['C' + str(random.randint(2, skinsCount))].value
    c = skins['C' + str(random.randint(2, skinsCount))].value
    d = skins['C' + str(random.randint(2, skinsCount))].value
    return [a, b, c, d]

def storeUpdate():
    ubd = openpyxl.load_workbook('users.xlsx')
    uinf = ubd['usersInfo']
    raw = 2
    while uinf['G' + str(raw)].value != None:
        uinf['G' + str(raw)].value = str(generateRandomStore())
        raw+=1
    ubd.save('users.xlsx')
    print('Магазин обновлён!')

async def on_startup():
        await db.db_start()
        print('дб запущен')

storeUpdate()
@disp.message(F.text == '/start')
async def cmdStart(message: types.Message):
    await db.cmdStart(message.from_user.id)
    db.cur.execute('USE usersINFO')
    db.cur.execute('INSERT INTO usersinfo VALUES ({key})')

    username = message.from_user.username
    regtime = f'{datetime.datetime.now().day}.{datetime.datetime.now().month}.{datetime.datetime.now().year}'  # дата рег-ции [17 02 2023]

    key = db.cur.execute('SELECT DISTINCT FROM usersInfo WHERE tgid == None')
    tgid = db.cur.execute('SELECT DISTINCT FROM usersInfo WHERE ')

    if message.from_user.username is None:
        username = message.from_user.first_name + str(key)
    if uinf[botidCell].value == message.from_user.id:  # если уже зарегестрирован
        await message.answer(f'Вы уже зарегистрированы как {uinf["A" + tgidCell[0][1]].value}')
    elif uinf[botidCell].value == None:  # регистрация


 #________________^^^НОВОЕ^^^_____________________
    #----------СТАРОЕ, ПЕРЕДЕЛАТЬ В SQL--------------
'''
        
        uinf[tgidCell[0]] = username # установка tgid
        uinf[botidCell] = message.from_user.id # установка botid формата [номер][первая буква]
        uinf[balanceCell] = 17000 # установка баланса
        uinf[regtimeCell] = regtime # установка даты рег-ии
        uinf[changingidCell] = 'none' # установка состояния сменяемого оружия
        uinf[storeCell] = str(generateRandomStore())

        uequip = bd['usersEquipped']
        botidEquip = findCell(uinf[tgidCell[0]].value, 'A', 1, 'users.xlsx', 'usersInfo')
        uequip[botidEquip[0]] = uinf[botidCell].value
        bd.save('users.xlsx') # сохранение бд
        setDefaultSkins(uinf[botidCell].value) # установка дефолтных скинов пользователю

        bd = openpyxl.load_workbook('users.xlsx')  # открываю бд
        uinf = bd['usersInfo']
        uequip = bd['usersEquipped']
        botidEquip = findCell(uinf[tgidCell[0]].value, 'A', 1, 'users.xlsx', 'usersInfo')

        fullinv = [] #ke = ueqip['B' + botIdCell[0][1]].value
        for i in range(1, 19):
            column = columns[i]
            fullinv.append(uequip[column + botidEquip[0][1]].value)
        uinf[fullinvCell].value = ', '.join(fullinv)
        bd.save('users.xlsx')

        await message.answer(f'Добро пожаловать в ValoShop!\nВы успешно зарегистрировались как {uinf[tgidCell[0]].value}\n\nПомощь - /help')
        await message.answer(f'Вам начислено 17000 VP', reply_markup=kb.main)
        print(f'{message.from_user.id} зарегистрировался в системе')
'''
@disp.message(lambda message: message.text in ['Профиль', 'Жопа'])
async def cmdProfile(message: types.Message):
    bd = openpyxl.load_workbook('users.xlsx')  # открываю бд
    uinf = bd['usersInfo']  # выбираю лист usersInfo
    botid = findCell(message.from_user.id, 'B', 2, 'users.xlsx', 'usersInfo') # поиск юзера в базе
    tgid = uinf['A' + botid[0][1]].value
    regtime = 'C' + botid[0][1] # ячейка для даты
    if botid[1] == True: # если юзер не зареган
        await message.reply('Вы ещё не зарегистрированы. Для регистрации введите /start')
    else: # если юзер найден
        stats = getUserStats(message.from_user.id)
        await message.reply(f'Профиль пользователя @{tgid}:\n\n'
                            f'Показатели:\n'
                            f'Здоровье: {stats[0]}\n'
                            f'Броня: {stats[1]}\n'
                            f'Урон: {stats[2]}\n'
                            f'Меткость: {stats[3]}%\n'
                            f'Шанс попадания в голову: {stats[4]}%\n'
                            f'Уклонение: {stats[5]}%\n\n'
                            f'Дата регистрации: {uinf[regtime].value}\n', reply_markup=kb.main)
        print(f'{message.from_user.id} запросил профиль')

@disp.message(F.text == 'Помощь')
async def cmdHelp(message: types.Message):
    await message.reply(f'Команды бота:\n\n/start - Регистрация\n/profile - Профиль\n/help - Помощь\n/inv - Инвентарь\n/store - Магазин\n{"_"*23}\nСоздатель: @magnitgd', reply_markup=kb.main)

@disp.message(F.text == 'Инвентарь')
async def cmdInv(message: types.Message):

    if isRegistered(message.from_user.id) == False:
        await message.reply('Вы ещё не зарегистрированы. Для регистрации введите /start')
        return
    tgidCell = findCell(message.from_user.id, 'B', 2, 'users.xlsx', 'usersInfo')  # формат переменной [координаты ячейки, содержимое]

    bd = openpyxl.load_workbook('users.xlsx')
    uinfo = bd['usersInfo']

    username = uinfo['A' + tgidCell[0][1]].value
    botIdCell = findCell(message.from_user.id, 'A', 2, 'users.xlsx', 'usersEquipped')

    ubd = openpyxl.load_workbook('users.xlsx')
    sbd = openpyxl.load_workbook('skins.xlsx')
    ueqip = ubd['usersEquipped']
    skins = sbd['skins']

    ke = ueqip['B' + botIdCell[0][1]].value
    raw = findCell(ke, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    kestring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # knife

    cc = ueqip['C' + botIdCell[0][1]].value
    raw = findCell(cc, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    ccstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # classic

    sy = ueqip['D' + botIdCell[0][1]].value
    raw = findCell(sy, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    systring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # shorty

    fy = ueqip['E' + botIdCell[0][1]].value
    raw = findCell(fy, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    fystring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # frenzy

    gt = ueqip['F' + botIdCell[0][1]].value
    raw = findCell(gt, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    gtstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # ghost

    sf = ueqip['G' + botIdCell[0][1]].value
    raw = findCell(sf, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    sfstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # sheriff

    by = ueqip['H' + botIdCell[0][1]].value
    raw = findCell(by, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    bystring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # bucky

    ml = ueqip['I' + botIdCell[0][1]].value
    raw = findCell(ml, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    mlstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # marshal

    sr = ueqip['J' + botIdCell[0][1]].value
    raw = findCell(sr, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    srstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # stinger

    ass = ueqip['K' + botIdCell[0][1]].value
    raw = findCell(ass, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    assstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # ares

    se = ueqip['L' + botIdCell[0][1]].value
    raw = findCell(se, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    sestring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # spectre

    je = ueqip['M' + botIdCell[0][1]].value
    raw = findCell(je, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    jestring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # judge

    bg = ueqip['N' + botIdCell[0][1]].value
    raw = findCell(bg, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    bgstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # bulldog

    gn = ueqip['O' + botIdCell[0][1]].value
    raw = findCell(gn, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    gnstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # guardian

    pm = ueqip['P' + botIdCell[0][1]].value
    raw = findCell(pm, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    pmstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # phantom

    vl = ueqip['Q' + botIdCell[0][1]].value
    raw = findCell(vl, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    vlstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # vandal

    on = ueqip['R' + botIdCell[0][1]].value
    raw = findCell(on, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    onstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] # odin

    orr = ueqip['S' + botIdCell[0][1]].value
    raw = findCell(orr, 'C', 1, 'skins.xlsx', 'skins')[0][1:]
    orrstring = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value] #opeator

    #skinChangeCmd = [kestring[1], ccstring[1], systring[1], fystring[1], gtstring[1], sfstring[1], bystring[1], mlstring[1], srstring[1], assstring[1], sestring[1], jestring[1], bgstring[1], gnstring[1], pmstring[1], vlstring[1], onstring[1], orrstring[1]]

    keText = '🔹' + f'{kestring[2]}  {kestring[1]} | {kestring[0]} здоровья'
    ccText = '🔹' +  f'{ccstring[2]}  {ccstring[1]} | {ccstring[0]} урона'
    syText =  '🔹' + f'{systring[2]}  {systring[1]} | {systring[0]}% шанс уклонения'
    fyText =  '🔹' + f'{fystring[2]}  {fystring[1]} | {fystring[0]} брони'
    gtText = '🔹' +  f'{gtstring[2]}  {gtstring[1]} | {gtstring[0]}% шанс попадания в голову'
    sfText =  '🔹' + f'{sfstring[2]}  {sfstring[1]} | {sfstring[0]}% меткости'
    byText = '🔹' +  f'{bystring[2]}  {bystring[1]} | шанс уклонения увеличен на {bystring[0]}%'
    mlText = '🔹' +  f'{mlstring[2]}  {mlstring[1]} | меткость увеличена на {mlstring[0]}%'
    srText = '🔹' +  f'{srstring[2]}  {srstring[1]} | к урону добавлено {srstring[0]} ед.'
    assText = '🔹' +  f'{assstring[2]}  {assstring[1]} | к броне добавлено {assstring[0]} ед.'
    seText =  '🔹' + f'{sestring[2]}  {sestring[1]} | шанс попадания в голову увеличен на {sestring[0]}%'
    jeText =  '🔹' + f'{jestring[2]}  {jestring[1]} | к уклонению добавлено {jestring[0]}%'
    bgText = '🔹' +  f'{bgstring[2]}  {bgstring[1]} | к здоровью добавлено {bgstring[0]} ед.'
    gnText =  '🔹' + f'{gnstring[2]}  {gnstring[1]} | здоровье увеличено на {gnstring[0]}%'
    pmText =  '🔹' + f'{pmstring[2]}  {pmstring[1]} | к шансу попадания в голову добавлено {pmstring[0]}%'
    vlText =  '🔹' + f'{vlstring[2]}  {vlstring[1]} | урон увеличен на {vlstring[0]}%'
    onText =  '🔹' + f'{onstring[2]}  {onstring[1]} | броня увеличена на {onstring[0]}%'
    orrText =  '🔹' + f'{orrstring[2]}  {orrstring[1]} | к меткости добавлено {orrstring[0]}%'

    kbInventory = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text='Назад')],
                    [KeyboardButton(text=keText)],
                    [KeyboardButton(text=ccText)],
                    [KeyboardButton(text=syText)],
                    [KeyboardButton(text=fyText)],
                    [KeyboardButton(text=gtText)],
                    [KeyboardButton(text=sfText)],
                    [KeyboardButton(text=byText)],
                    [KeyboardButton(text=mlText)],
                    [KeyboardButton(text=srText)],
                    [KeyboardButton(text=assText)],
                    [KeyboardButton(text=seText)],
                    [KeyboardButton(text=jeText)],
                    [KeyboardButton(text=bgText)],
                    [KeyboardButton(text=gnText)],
                    [KeyboardButton(text=pmText)],
                    [KeyboardButton(text=vlText)],
                    [KeyboardButton(text=onText)],
                    [KeyboardButton(text=orrText)]],
                    resize_keyboard=True, input_field_placeholder='Нажмите для смены скина')
    await message.answer(f'Инвентарь {username}:', reply_markup=kbInventory)
    print(f'{message.from_user.id} открыл инвентарь')

@disp.message(lambda message: message.text.split(' | ')[0][1:] in allSkins)
async def cmdChangeSkin(message: Message):
    # Определение оружия
    skinType = message.text.split(' | ')[0][1:].split('  ')[1]  # строка с названием гана
    sType = skinType
    ubd = openpyxl.load_workbook('users.xlsx')
    sbd = openpyxl.load_workbook('skins.xlsx')
    uinfo = ubd['usersInfo']
    skins = sbd['skins']
    a = findCell(str(skinType), 'A', 2, 'skins.xlsx', 'skins')  # поиск
    a = str(a[0][0] + str(int(a[0][1:]) + 1))  # и преобразование в формат [A1]
    skinType = skins['C' + a[1:]].value[:2]  # определение типа ID формата 'ke'

    # Обновление состояния смены оружия
    a = findCell(message.from_user.id, 'B', 1, 'users.xlsx', 'usersInfo')
    a = 'F' + a[0][1:]
    uinfo[a].value = skinType
    ubd.save('users.xlsx')

    # Поиск скинов на это оружие
    uid = message.from_user.id
    a = findCell(uid, 'B', 1, 'users.xlsx', 'usersInfo')
    a = a[0]
    uinv = uinfo['E' + a[1:]].value
    uinv = uinv.split(', ')  # получение инвентаря пользователя в виде списка
    uskins = []  # список для скинов конкретного типа
    kbSkinChange = ReplyKeyboardBuilder()
    kbSkinChange.button(text='Назад')

    text = f'Ваши скины на {sType}:'
    for i in range(0, len(uinv)):  # пополняет список скинов типа
        if skinType in uinv[i]:
            uskins.append(uinv[i])
    for i in range(0, len(uskins)):  # Формирование текста сообщения
        raw = findCell(uskins[i], 'C', 1, 'skins.xlsx', 'skins')[0][1:]
        skin = [skins['D' + raw].value, skins['A' + raw].value, skins['B' + raw].value]  # [effect, type, skin name]
        if skin[1] == skin[2]:
            kbSkinChange.button(text=f'{skin[2]} | Эффект: {skin[0]}')
        else:
            kbSkinChange.button(text=f'{skin[2]}  {skin[1]} | Эффект: {skin[0]}')

    await message.answer(text, reply_markup=kbSkinChange.as_markup(), input_field_placeholder='Выберите желаемый скин')
    ubd.save('users.xlsx')
    sbd.save('skins.xlsx')
    print(f'{message.from_user.id} решил сменить скин на {message.text[1:]}')

@disp.message(lambda message: message.text.split(' | ')[0] in allSkins)
async def cmdChooseSkin(message: Message):
    ubd = openpyxl.load_workbook('users.xlsx')
    sbd = openpyxl.load_workbook('skins.xlsx')
    uinfo = ubd['usersInfo']
    ueqip = ubd['usersEquipped']
    skins = sbd['skins']

    skinType = findCell(message.from_user.id, 'B', 1, 'users.xlsx', 'usersInfo')[0]
    skinType = uinfo['F' + skinType[1:]].value
    skin = message.text.split(' | ')[0].split('  ')[1]  # название коллекции
    a = findCell(skin, 'B', 1, 'skins.xlsx', 'skins')
    currentSkinType = skins['C' + a[0][1:]].value[:2]
    row = int(a[0][1:])
    while currentSkinType != skinType:
        row += 1
        a = findCell(skin, 'B', row, 'skins.xlsx', 'skins')
        currentSkinType = skins['C' + a[0][1:]].value[:2]
        row = int(a[0][1:])
    print(a, row, currentSkinType)
    skinid = skins['C' + a[0][1:]].value  # ke24
    # Поиск скинов на это оружие
    uid = message.from_user.id
    a = findCell(uid, 'B', 1, 'users.xlsx', 'usersInfo')
    a = a[0]
    uinv = uinfo['E' + a[1:]].value
    uinv = uinv.split(', ')  # получение инвентаря пользователя в виде списка
    if skinid in uinv:
        for i in range(0, len(userEquipCells)):
            if skinType == userEquipCells[i][1:]:
                a = findCell(message.from_user.id, 'B', 1, 'users.xlsx', 'usersInfo')[0]
                ueqip[userEquipCells[i][0] + a[1:]].value = skinid
                uinfo['F' + a[1:]].value = 'none'
                ubd.save('users.xlsx')
                await message.answer('Вы успешно сменили скин')
                print(f'{message.from_user.id} сменил скин')

    else:
        await message.answer('Скин отсутствует в вашей коллекции, либо возникла какая-то ошибка')
        ubd.save('users.xlsx')

async def main():
    await on_startup()
    await disp.start_polling(bot)



if __name__ == "__main__":
    asyncio.run(main())